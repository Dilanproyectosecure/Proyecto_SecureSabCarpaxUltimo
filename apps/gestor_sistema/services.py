import random
import string

from apps.login.models import Usuarios, Roles, RoleUser
from apps.reporte_monitoreo.coordinador.models import Ficha, AsistenciaSede
from django.utils import timezone
from .models import Huella, registro_actividad


def registrar_actividad(usuario, tipo_accion, actividad, descripcion="", request=None):
    try:
        ip_address = None
        user_agent = None

        ahora = timezone.localtime(timezone.now())

        if request:
            ip_address = request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR', ''))
            user_agent = request.META.get('HTTP_USER_AGENT', '')

        registro_actividad.objects.create(
            id_usuario=usuario,
            tipo_accion=tipo_accion,
            actividad=actividad,
            descripcion=descripcion,
            ip_address=ip_address,
            user_agent=user_agent,
            fecha=ahora.date(),
            hora=ahora.time(),
        )

    except Exception as e:
        print(f"Error al registrar actividad: {e}")

def registrar_asistencia_sede_por_huella(usuario, request=None, momento=None):
    momento = momento or timezone.localtime()
    hoy = momento.date()

    asistencia_hoy = AsistenciaSede.objects.filter(
        id_usuario=usuario,
        fecha=hoy,
    ).order_by('-id_asistencia').first()

    if not asistencia_hoy:
        asistencia = AsistenciaSede.objects.create(
            id_usuario=usuario,
            fecha=hoy,
            hora_entrada=momento.time(),
            estado_asistencia='Entrada',
        )

        registrar_actividad(
            usuario=usuario,
            tipo_accion='HUELLA_SEDE_ENTRADA',
            actividad='Entrada en sede por huella',
            descripcion=(
                f'Entrada registrada por huella para {usuario.nombre} {usuario.apellido} '
                f'(cédula: {usuario.cedula}) a las {momento:%H:%M:%S}'
            ),
            request=request,
        )

        return {
            'estado': 'entrada',
            'asistencia': asistencia,
        }

    if not asistencia_hoy.hora_salida:
        asistencia_hoy.hora_salida = momento.time()
        asistencia_hoy.estado_asistencia = 'Salida'
        asistencia_hoy.save(update_fields=['hora_salida', 'estado_asistencia'])

        registrar_actividad(
            usuario=usuario,
            tipo_accion='HUELLA_SEDE_SALIDA',
            actividad='Salida de sede por huella',
            descripcion=(
                f'Salida registrada por huella para {usuario.nombre} {usuario.apellido} '
                f'(cédula: {usuario.cedula}) a las {momento:%H:%M:%S}'
            ),
            request=request,
        )

        return {
            'estado': 'salida',
            'asistencia': asistencia_hoy,
        }

    # Ya tiene entrada y salida → es un reingreso, crear nuevo registro
    asistencia = AsistenciaSede.objects.create(
        id_usuario=usuario,
        fecha=hoy,
        hora_entrada=momento.time(),
        estado_asistencia='Entrada',
    )

    registrar_actividad(
        usuario=usuario,
        tipo_accion='HUELLA_SEDE_ENTRADA',
        actividad='Reingreso en sede por huella',
        descripcion=(
            f'Reingreso registrado por huella para {usuario.nombre} {usuario.apellido} '
            f'(cédula: {usuario.cedula}) a las {momento:%H:%M:%S}'
        ),
        request=request,
    )

    return {
        'estado': 'entrada',
        'asistencia': asistencia,
    }


def crear_usuario(request, datos):
    cedula = datos.get('cedula')
    nombre = datos.get('nombre')
    apellido = datos.get('apellido')
    correo = datos.get('correo')
    telefono = datos.get('telefono')
    password = datos.get('password') or generar_password_segura(datos.get('cedula', ''), datos.get('nombre', ''))
    rol_id = datos.get('rol_id')
    ficha_id = datos.get('ficha_id')

    rol_nombre = 'sin rol'
    rol = None
    if rol_id:
        try:
            rol = Roles.objects.get(id=rol_id)
            rol_nombre = rol.name
        except Roles.DoesNotExist:
            rol_nombre = 'rol no encontrado'

    usuario = Usuarios.objects.create_user(
        cedula=cedula,
        password=password,
        nombre=nombre,
        apellido=apellido,
        correo=correo,
        telefono=telefono,
        is_active=True,
        is_staff=(rol_nombre == 'gestor'),
        estado='Activo',
    )

    if rol:
        RoleUser.objects.update_or_create(id_usuario=usuario, defaults={'role': rol})

    if ficha_id:
        usuario.id_ficha_id = ficha_id
        usuario.save()

    registrar_actividad(
        usuario=request.user,
        tipo_accion='CREATE',
        actividad='Creación de usuario',
        descripcion=f'Se creó el usuario {nombre} {apellido} con cédula {cedula} y rol {rol_nombre}',
        request=request,
    )

    return usuario


def actualizar_usuario(usuario, datos):
    usuario.nombre = datos.get('nombre', usuario.nombre)
    usuario.apellido = datos.get('apellido', usuario.apellido)
    usuario.correo = datos.get('correo', usuario.correo)
    usuario.telefono = datos.get('telefono', usuario.telefono)
    usuario.save()

    rol_id = datos.get('rol_id')
    if rol_id:
        rol = Roles.objects.filter(id=rol_id).first()
        if rol:
            RoleUser.objects.update_or_create(id_usuario=usuario, defaults={'role': rol})

    return usuario


def eliminar_usuario(request, usuario):
    registrar_actividad(
        usuario=request.user,
        tipo_accion='DELETE',
        actividad='Eliminación de usuario',
        descripcion=f'Se eliminó al usuario {usuario.nombre} {usuario.apellido} con cédula {usuario.cedula}',
        request=request,
    )
    usuario.delete()


def cambiar_estado_usuario(request, usuario, accion):
    if accion == 'desactivar':
        usuario.estado = 'Inactivo'
        usuario.is_active = False
        tipo_accion = 'DESACTIVAR'
        actividad = 'Desactivación de usuario'
        descripcion = f'Se desactivó al usuario {usuario.nombre} {usuario.apellido} (cédula: {usuario.cedula})'
    else:
        usuario.estado = 'Activo'
        usuario.is_active = True
        tipo_accion = 'ACTIVAR'
        actividad = 'Activación de usuario'
        descripcion = f'Se activó al usuario {usuario.nombre} {usuario.apellido} (cédula: {usuario.cedula})'

    usuario.save()

    registrar_actividad(
        usuario=request.user,
        tipo_accion=tipo_accion,
        actividad=actividad,
        descripcion=descripcion,
        request=request,
    )


def procesar_carga_masiva(request, archivo):
    import csv
    from .views import generar_password_segura, enviar_password_usuario

    filename = getattr(archivo, 'name', '')

    if filename.endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(archivo, read_only=True)
        ws = wb.active
        raw_headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(max_row=1))]
        rows_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, raw_h in enumerate(raw_headers):
                if raw_h:
                    val = row[i] if i < len(row) else None
                    row_dict[raw_h] = str(val).strip() if val is not None else ''
            rows_data.append(row_dict)
    elif filename.endswith('.csv'):
        content = archivo.read().decode('utf-8')
        if content.startswith('\ufeff'):
            content = content[1:]
        reader = csv.DictReader(content.splitlines())
        rows_data = []
        for row in reader:
            cleaned = {}
            for k, v in row.items():
                if k:
                    cleaned[k.strip().lower()] = str(v).strip() if v else ''
            rows_data.append(cleaned)
    else:
        return 0, 0, ['Formato no soportado. Use .csv o .xlsx']

    creados = 0
    omitidos = 0
    errores = []

    for index, row in enumerate(rows_data, start=2):
        cedula = row.get('cedula', '').strip()
        nombre = row.get('nombre', '').strip()
        apellido = row.get('apellido', '').strip()
        correo = row.get('correo', '').strip()
        telefono = row.get('telefono', '').strip()
        rol_nombre = row.get('rol', '').strip().lower()
        ficha_numero = row.get('ficha', '').strip()
        tipo_doc = row.get('tipo_documento', '').strip()

        if not cedula or not nombre:
            errores.append(f'Fila {index}: Faltan campos obligatorios (cedula, nombre)')
            continue

        if Usuarios.objects.filter(cedula=cedula).exists():
            omitidos += 1
            continue

        rol = Roles.objects.filter(name=rol_nombre).first()
        if not rol:
            errores.append(f'Fila {index}: Rol "{rol_nombre}" no existe')
            continue

        password_generada = generar_password_segura(cedula, nombre)

        usuario = Usuarios.objects.create_user(
            cedula=cedula,
            password=password_generada,
            nombre=nombre,
            apellido=apellido,
            correo=correo,
            telefono=telefono,
            tipo_documento=tipo_doc or 'CC',
            is_active=True,
            is_staff=(rol_nombre == 'gestor'),
            estado='Activo',
        )

        RoleUser.objects.update_or_create(id_usuario=usuario, defaults={'role': rol})

        if rol_nombre == 'aprendiz' and ficha_numero:
            ficha = Ficha.objects.filter(numero_ficha=ficha_numero).first()
            if ficha:
                usuario.id_ficha = ficha
                usuario.save()

        if correo:
            enviar_password_usuario(nombre, apellido, correo, password_generada, cedula)

        creados += 1

    return creados, omitidos, errores