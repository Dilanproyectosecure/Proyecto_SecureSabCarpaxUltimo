import openpyxl
from django.core.management.base import BaseCommand
from django.db import connections


class Command(BaseCommand):
    help = 'Sincroniza competencias y resultados de aprendizaje desde Excel a la BD'

    def add_arguments(self, parser):
        parser.add_argument('archivo', type=str, help='Ruta al archivo Excel')

    def handle(self, *args, **options):
        archivo = options['archivo']
        wb = openpyxl.load_workbook(archivo, read_only=True)
        ws = wb.active

        raw_headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(max_row=1))]

        competencias_excel = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(raw_headers, row))
            nombre_comp = str(data.get('competencia', '') or '').strip()
            nombre_ra = str(data.get('resultado de aprendizaje', '') or '').strip()
            juicio = str(data.get('juicio a evaluar', '') or '').strip()
            if nombre_comp and nombre_ra:
                if nombre_comp not in competencias_excel:
                    competencias_excel[nombre_comp] = []
                competencias_excel[nombre_comp].append({
                    'ra': nombre_ra,
                    'juicio': juicio,
                })

        self.stdout.write(f'Excel: {len(competencias_excel)} competencias, {sum(len(v) for v in competencias_excel.values())} RA')

        conn = connections['default']
        cursor = conn.cursor()

        cursor.execute('SELECT id_competencia, nombre_competencia FROM competencia')
        bd_competencias = {row[1].strip(): row[0] for row in cursor.fetchall()}
        self.stdout.write(f'BD actual: {len(bd_competencias)} competencias')

        comp_creadas = 0
        ra_creados = 0
        ra_existentes = 0

        for nombre_comp, ras in competencias_excel.items():
            if nombre_comp in bd_competencias:
                comp_id = bd_competencias[nombre_comp]
            else:
                cursor.execute(
                    'INSERT INTO competencia (nombre_competencia, id_programa, estado) VALUES (%s, 2, %s)',
                    [nombre_comp, 'Activa']
                )
                comp_id = cursor.lastrowid
                bd_competencias[nombre_comp] = comp_id
                comp_creadas += 1
                self.stdout.write(f'  + Competencia: {nombre_comp[:70]} (id={comp_id})')

            for ra_data in ras:
                ra_texto = ra_data['ra']

                cursor.execute(
                    'SELECT id_Resultado_Aprendizaje FROM resultado_aprendizaje WHERE Resultado_Aprendizaje = %s AND id_competencia = %s',
                    [ra_texto, comp_id]
                )
                existente = cursor.fetchone()

                if existente:
                    ra_existentes += 1
                else:
                    cursor.execute(
                        'INSERT INTO resultado_aprendizaje (Resultado_Aprendizaje, id_competencia, trimestre) VALUES (%s, %s, 1)',
                        [ra_texto, comp_id]
                    )
                    ra_creados += 1

        conn.commit()

        self.stdout.write(self.style.SUCCESS(f'\n=== RESUMEN ==='))
        self.stdout.write(f'Competencias nuevas creadas: {comp_creadas}')
        self.stdout.write(f'Resultados de aprendizaje creados: {ra_creados}')
        self.stdout.write(f'Resultados de aprendizaje ya existentes: {ra_existentes}')
        self.stdout.write(self.style.SUCCESS(f'Total competencias en BD: {len(bd_competencias)}'))
